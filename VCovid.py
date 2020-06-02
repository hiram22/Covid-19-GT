# -*- coding: utf-8 -*-
"""
Created on Tue May 19 11:08:03 2020

@author: Steven Rubio
"""


import pandas as pd
from pandas import ExcelFile
import matplotlib.pyplot as plt
from rich import print
from rich.console import Console
from rich.progress import track
from openpyxl import load_workbook
from openpyxl import Workbook
    
#Consola para imprimir mensajes en pantalla. 
console = Console()
console.print()

#   Datos descargados de:
#   https://www.mspas.gob.gt/index.php/noticias/covid-19/casos
df = pd.ExcelFile('infocovid19.xlsx')

#   Funcion para generar el tiempo que a pasado para duplicar los casos
def DuplicaC(List, Vo):
    # ----------
    # Entradas:
    # List      = Lista con la cantidad de casos
    # Vo        = Valor inicial
    # ----------
    # Salidas:
    # casos     = Lista con el número de casos
    # tiempo    = Lista con el tiempo para cada número de casos
    # ultimo    = Siguiente número de casos a alcanzar
    # contador  = Tiempo actual en el número de casos
    # ----------
    casos = [(Vo)]
    tiempo = [0]
    contador = 0
    longl = len(List)
    
    for i in track(range(longl)):
        ultimo =  2*casos[-1]
        if(List[i]>=ultimo):
            tiempo.append((contador))
            casos.append(ultimo)
            contador = 0
        else:
            if(List[i]!=0):    contador=contador+1
        i
    #Print fancy en consola en caso de no ver gráficas.
    #Es una forma rápida de saber los últimos
    console.print()
    console.print('Siguiente valor: '+str(ultimo)+'. Días hasta el momento: '+str(contador)+'.')
    style ,style2= 'yellow','magenta'
    console.print('Valor actual: [bold '+style+']'+str(List[len(List)-1])+'[/bold '+style+'].')
    console.print('Último tiempo de duplicación: [bold '+style2+']'+str(tiempo[len(tiempo)-1])+'[/bold '+style2+'].')
    console.print('*-------------------------------------------------*')
    console.print()
    
    return casos,tiempo,ultimo,contador
    

# Función para obtener los casos diarios Regionales
def DiarioR(Lista):
        # ----------
    # Entradas:
    # Lista      = Lista con la cantidad de casos Regionales
    # ----------
    # Salidas:
    # NuevaLista = Lista con el número de casos por día
    # ----------
    NuevaLista = []
    for i in range(1,len(Lista)):
        val = Lista[i]-Lista[i-1]
        NuevaLista.append(val)
        
    return NuevaLista
    
#Obtenemos última fecha registrada
df0 = pd.ExcelFile('Myinfocovid19.xlsx')
S0 = df0.parse('Casos por región')
f2 = S0['Fecha']

#Obtenemos la última fecha de nuestros datos locales
SL = df.parse('Casos por día')
f = SL['Fecha']

if(f2[len(f2)-1]==f[len(f)-1]):
    print("Actualizado")
else:
    print('No Actualizado')
    fechaN = f[len(f)-1]

    CR = df.parse('Casos por región')
    R1 = CR.columns[2]
    R2 = CR[R1][0]
    R3 = CR[R1][1]
    R4 = CR[R1][2]
    R5 = CR[R1][3]
    RT = R1+R2+R3+R4+R5
    newRow = (fechaN,R1,R2,R3,R4,R5,RT)
    dfn = pd.DataFrame(newRow)
    dfn = dfn.transpose()
    #dfn.to_excel('Myinfocovid19.xlsx',sheet_name='Casos por región')

    wb = load_workbook(filename = 'Myinfocovid19.xlsx', read_only=False)
    #Get the current Active Sheet
    ws = wb.active
    #Columna a escribir
    COL = len(S0)+2
    
    print(COL)
    
    #Escribimos nuevos valores
    ws.cell(COL,1,value = fechaN)
    ws.cell(COL,2,value = R1)
    ws.cell(COL,3,value = R2)
    ws.cell(COL,4,value = R3)
    ws.cell(COL,5,value = R4)
    ws.cell(COL,6,value = R5)
    ws.cell(COL,7,value = RT)
    print("Intentando Guardar")
    wb.save(filename = 'Myinfocovid19.xlsx')

#Ajuste para que se muestre en formado dd:mm:yy
fecha = str(f[len(f)-1])
fecha = fecha[:10]
fecha = fecha[8:10]+fecha[4:8]+fecha[:4]

#Creacion del plot 1. 
fig1 , ax = plt.subplots(1,2,figsize=(14, 5))

#Modifico los fonts del plot
#Opciones con: print(plt.style.available)
plt.style.use('bmh')

#Plot the data
#Primer plot
S1 = df.parse('evolución de casos')

x = S1['Casos recuperados']
y = S1['Casos activos']
z = S1['Casos fallecidos']

console.print('Casos Activos >>>', style="bold Blue")
val , tmp,  ult,  cnt   = DuplicaC(y,1)
console.print('Casos Fallecidos >>>', style="bold Red")
val2, tmp2, ult2, cnt2  = DuplicaC(z,2)
console.print('Casos Recuperados >>>', style="bold Green")
val3, tmp3, ult3, cnt3  = DuplicaC(x,4)

#ax[0].set_ylim([0, 20])
ax[1].semilogx(val, tmp, label='Casos Activos',color='#C725E1')
ax[1].semilogx(val2, tmp2, label='Casos Fallecidos',color='#25D8E1')
ax[1].semilogx(val3, tmp3, label='Casos Recuperados',color='g')

ax[1].semilogx((val[-1],ult),(tmp[-1],cnt),'o--',color='#C725E1')
ax[1].annotate('Valor actual', xy=((ult, cnt)), xytext=((ult-2000, cnt-2)),
              arrowprops=dict(facecolor='black', shrink=0.05))
ax[1].semilogx((val2[-1],ult2),(tmp2[-1],cnt2),'o--',color='#25D8E1')
ax[1].semilogx((val3[-1],ult3),(tmp3[-1],cnt3),'go--')

#Nombre
ax[1].set_title('Duplicación de Casos Covid-19 Guatemala ('+fecha+')')
ax[1].set(xlabel='Total de Casos',ylabel='Días')

# Add a legend
ax[1].legend()

#Datos
df = pd.ExcelFile('infocovid19.xlsx')
S2 = df.parse('Casos por día')
#Segundo plot
x = S2['Casos por día']
z = S2['Casos recuperados']
y = S2['Casos fallecidos']


ax[0].plot(range(len(x)), x,label='Casos Nuevos')
ax[0].plot(range(len(y)), y,label='Casos Fallecidos',color='#25D8E1')
ax[0].plot(range(len(z)), z,label='Casos Recuperados',color='g')

#Nombre
ax[0].set_title('Casos por día Covid-19 Guatemala ('+fecha+')')
ax[0].set(xlabel='Días desde el caso 1 (13-03-2020)',ylabel='Casos')

# Add a legend
ax[0].legend(loc='upper left')

#Guardando el plot
#plt.savefig('C:/Users/HRV/Desktop/Post-U/Scripts/Covid-19-GT/boom.png')


#Segundo plot
fig2 , ax2 = plt.subplots(1,2,figsize=(14, 5))

#Plot A
df2 = pd.ExcelFile('Myinfocovid19.xlsx')
S2 = df2.parse('Casos por región')

#Data
r1 = S2['Region 1']
r2 = S2['Region 2']
r3 = S2['Region 3']
r4 = S2['Region 4']
r5 = S2['Region 5']
f2 = S2['Fecha']
f_i = range(len(f2))

ax2[0].semilogy(f_i, r1,label='Región 1',color = '#FB1D07')
ax2[0].semilogy(f_i, r2,label='Región 2',color = '#0A9B11')
ax2[0].semilogy(f_i, r3,label='Región 3',color = '#15378F')
ax2[0].semilogy(f_i, r4,label='Región 4',color = '#8D2294')
ax2[0].semilogy(f_i, r5,label='Región 5',color = 'yellow')

#Nombre
ax2[0].set_title('Total de casos por región')
ax2[0].set(ylabel='Cantidad de Casos',xlabel='Días a partir del 13-04-2020 ') 

#Limito el número de valores que se muesttran en el eje y       
plt.locator_params(axis='y', nbins=4)

# Add a legend
ax2[0].legend()

#Plot B
R1D = DiarioR(r1)
R2D = DiarioR(r2)
R3D = DiarioR(r3)
R4D = DiarioR(r4)
R5D = DiarioR(r5)

ax2[1].plot(range(len(R1D)),R1D,label='Región 1',color = '#FB1D07')
ax2[1].plot(range(len(R2D)),R2D,label='Región 2',color = '#0A9B11')
ax2[1].plot(range(len(R3D)),R3D,label='Región 3',color = '#15378F')
ax2[1].plot(range(len(R4D)),R4D,label='Región 4',color = '#8D2294')
ax2[1].plot(range(len(R5D)),R5D,label='Región 5',color = 'yellow')

#Nombre
ax2[1].set_title('Casos por día por región')
ax2[1].set(ylabel='Cantidad de Casos',xlabel='Días a partir del 14-04-2020 ') 

# Add a legend
ax2[1].legend()

#Tercer Plot
fig3 , ax3 = plt.subplots(1,1,figsize=(10, 10))

#Leemos los datos
S1 = df.parse('evolución de casos')
f3 = S1['Fecha']
x = S1['Casos recuperados']
y = S1['Casos activos']
z = S1['Casos fallecidos']

#Ploteamos en escala normal y logy
ax3.plot(range(len(y)), y,label='Casos Activos',color = '#15378F')
#ax3[1].semilogy(range(len(y)), y,label='Casos Activos',color = '')
ax3.plot(range(len(x)), x,label='Casos Recuperados',color = '#07E0D6')
#ax3[1].semilogy(range(len(x)), x,label='Casos Recuperados',color = '#07E0D6')
ax3.plot(range(len(z)), z,label='Casos Fallecidos',color = '#FB1D07')
#ax3[1].semilogy(range(len(z)), z,label='Casos Fallecidos',color = '#FB1D07')

#Lables
ax3.set_title('Casos acumulados Covid-19 Guatemala ('+fecha+')')
ax3.set(xlabel='Días desde el caso 1 (13-03-2020)',ylabel='Casos')

# Add a legend
ax3.legend()

#Show the plot
plt.show()

#Guardando el plot
#plt.savefig('C:/Users/HRV/Desktop/Post-U/Scripts/Covid-19-GT/boom2.png')

